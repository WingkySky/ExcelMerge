import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import schedule
import time
import os
from datetime import datetime
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from ttkthemes import ThemedStyle

class ExcelMergerApp:
    def configure_styles(self):
        """配置所有自定义样式"""
        # 设置全局字体
        default_font = ('Microsoft YaHei UI', 10)  # 使用微软雅黑作为默认字体
        self.style.configure('.', font=default_font)
        
        # 自定义按钮样式
        self.style.configure('Custom.TButton', 
                           padding=5, 
                           font=default_font)
        
        # 自定义标签样式
        self.style.configure('Custom.TLabel', 
                           font=default_font,
                           padding=2)
        
        # 自定义框架样式
        self.style.configure('Custom.TLabelframe', 
                           font=default_font,
                           padding=5)
        
        self.style.configure('Custom.TLabelframe.Label', 
                           font=('Microsoft YaHei UI', 10, 'bold'))
        
        # 自定义输入框样式
        self.style.configure('Custom.TEntry',
                           padding=5)
        
        # 自定义Treeview样式
        self.style.configure('Custom.Treeview',
                           font=default_font,
                           rowheight=25)
        
        self.style.configure('Custom.Treeview.Heading',
                           font=('Microsoft YaHei UI', 10, 'bold'),
                           padding=5)
        
        # 自定义Checkbutton样式
        self.style.configure('Custom.TCheckbutton',
                           font=default_font,
                           padding=5)
        
        # 自定义Radiobutton样式
        self.style.configure('Custom.TRadiobutton',
                           font=default_font,
                           padding=5)
        
        # 自定义Combobox样式
        self.style.configure('Custom.TCombobox',
                           padding=5,
                           font=default_font)

    def __init__(self, root):
        self.root = root
        self.root.title("Excel文件合并工具")
        
        # 设置主题和样式
        self.style = ThemedStyle(self.root)
        self.style.set_theme("clearlooks")  # 使用现代化的cleanlook主题
        
        # 配置自定义样式
        self.configure_styles()
        
        # 获取屏幕尺寸
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # 设置固定的窗口大小
        window_width = 600  # 固定宽度
        window_height = 700  # 固定高度
        
        # 计算窗口位置（居中）
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口大小和位置
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 设置最小窗口大小
        self.root.minsize(600, 700)  # 与固定大小相同        
        
        
        # 存储选择的文件和输出路径
        self.input_files = []
        self.output_path = ""
        self.output_path_var = tk.StringVar()  # 添加输出路径变量
        self.output_filename_var = tk.StringVar(value="合并文件")  # 添加输出文件名变量
        
        # 默认路径列表
        self.default_paths = [
            os.path.expanduser("~/Documents"),  # 文档文件夹
            os.path.expanduser("~/Desktop"),    # 桌面
            os.path.dirname(os.path.abspath(__file__)),  # 程序所在目录
        ]
        # 尝试添加最近使用的路径
        self.recent_paths_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "recent_paths.txt")
        self.recent_paths = self.load_recent_paths()
        
        # 合并默认路径和最近路径
        self.available_paths = self.recent_paths + [p for p in self.default_paths if p not in self.recent_paths]
        
        # 如果有可用路径，设置第一个为默认路径
        if self.available_paths:
            self.output_path = self.available_paths[0]
            self.output_path_var.set(self.available_paths[0])

        self.schedule_time = "00:00"  # 默认定时时间
        self.is_scheduling = False
        
        # 存储文件的sheet信息
        self.file_sheets = {}  # {文件路径: [sheet名称列表]}
        self.selected_sheets = {}  # {文件路径: 选中的sheet名称}
        self.sheet_names = {}  # {文件路径: 自定义sheet名称}
        
        # 存储样式信息
        self.style_templates = {}  # {列名: 样式信息}
        
        # 合并设置
        self.merge_mode = tk.StringVar(value="single")  # single: 合并到单个sheet, multiple: 每个文件一个sheet
        self.sheet_name_mode = tk.StringVar(value="auto")  # auto: 使用文件名, original: 使用原sheet名, custom: 使用自定义名称
        self.custom_sheet_name = tk.StringVar(value="Sheet1")  # 自定义sheet名称
        
        # 存储数据区间信息
        self.start_row = tk.StringVar(value="1")
        self.end_row = tk.StringVar(value="")
        self.start_col = tk.StringVar(value="A")
        self.end_col = tk.StringVar(value="")
        
        # 表头设置
        self.header_row = tk.StringVar(value="1")  # 表头行号
        self.keep_header = tk.BooleanVar(value=True)  # 是否保留表头
        
        # 样式设置
        self.keep_styles = tk.BooleanVar(value=True)  # 是否保留样式
        self.keep_column_width = tk.BooleanVar(value=True)  # 是否保留列宽
        self.keep_cell_format = tk.BooleanVar(value=True)  # 是否保留单元格格式（数字、日期等）
        self.keep_colors = tk.BooleanVar(value=True)  # 是否保留颜色（背景色、字体色）
        
        # 状态变量
        self.status_var = tk.StringVar(value="就绪")  # 添加状态变量
        
        # 定时任务变量
        self.time_var = tk.StringVar(value=self.schedule_time)  # 添加时间变量
        
        self.create_gui()
        
    def create_gui(self):
        # 创建主框架，添加内边距和样式
        main_frame = ttk.Frame(self.root, padding="10", style='Custom.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建Notebook，设置样式
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建各个标签页
        file_page = ttk.Frame(notebook, style='Custom.TFrame')
        merge_page = ttk.Frame(notebook, style='Custom.TFrame')
        style_page = ttk.Frame(notebook, style='Custom.TFrame')
        schedule_page = ttk.Frame(notebook, style='Custom.TFrame')
        
        notebook.add(file_page, text="文件选择", padding=5)
        notebook.add(merge_page, text="合并设置", padding=5)
        notebook.add(style_page, text="样式设置", padding=5)
        notebook.add(schedule_page, text="定时任务", padding=5)
        
        # ===== 文件选择页面 =====
        # 文件和Sheet选择区域
        file_frame = ttk.LabelFrame(file_page, text="文件和Sheet选择", padding="10", style='Custom.TLabelframe')
        file_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 创建表格来显示文件和对应的sheet选择
        # 创建带滚动条的框架
        tree_frame = ttk.Frame(file_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("文件名", "选择Sheet", "自定义Sheet名")
        self.file_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8, style='Custom.Treeview')
        
        # 添加垂直滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.file_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 添加水平滚动条
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.file_tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 配置树形视图的滚动
        self.file_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 设置列标题
        for col in columns:
            self.file_tree.heading(col, text=col)
        
        # 设置合理的初始列宽
        self.file_tree.column("文件名", width=100)
        self.file_tree.column("选择Sheet", width=60)
        self.file_tree.column("自定义Sheet名", width=100)
        
        # 添加和清除按钮
        btn_frame = ttk.Frame(file_frame, style='Custom.TFrame')
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(btn_frame, text="添加文件", command=self.add_files, style='Custom.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清除所有", command=self.clear_files, style='Custom.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="修改选中文件的Sheet", command=self.change_sheet, style='Custom.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="修改Sheet名称", command=self.change_sheet_name, style='Custom.TButton').pack(side=tk.LEFT, padx=5)
        
        # 输出路径选择
        output_frame = ttk.LabelFrame(file_page, text="输出设置", padding="10", style='Custom.TLabelframe')
        output_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 默认路径选择
        path_frame = ttk.Frame(output_frame, style='Custom.TFrame')
        path_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(path_frame, text="快速路径：", width=10, style='Custom.TLabel').pack(side=tk.LEFT)
        
        # 创建下拉菜单
        self.path_combobox = ttk.Combobox(path_frame, values=[p for p in self.available_paths], width=45, style='Custom.TCombobox')
        if self.available_paths:
            self.path_combobox.set(self.available_paths[0])
        self.path_combobox.pack(side=tk.LEFT, padx=5)
        
        # 自定义路径
        custom_path_frame = ttk.Frame(output_frame, style='Custom.TFrame')
        custom_path_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(custom_path_frame, text="自定义路径：", width=10, style='Custom.TLabel').pack(side=tk.LEFT)
        ttk.Entry(custom_path_frame, textvariable=self.output_path_var, width=45, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        ttk.Button(custom_path_frame, text="浏览", command=self.select_output_path, style='Custom.TButton').pack(side=tk.LEFT)
        
        # 文件名设置
        filename_frame = ttk.Frame(output_frame, style='Custom.TFrame')
        filename_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(filename_frame, text="文件名：", width=10, style='Custom.TLabel').pack(side=tk.LEFT)
        ttk.Entry(filename_frame, textvariable=self.output_filename_var, width=30, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        ttk.Label(filename_frame, text=".xlsx", style='Custom.TLabel').pack(side=tk.LEFT)
        
        # ===== 合并设置页面 =====
        # 合并方式设置
        merge_settings_frame = ttk.LabelFrame(merge_page, text="合并方式设置", padding="10", style='Custom.TLabelframe')
        merge_settings_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 合并方式选择
        merge_mode_frame = ttk.Frame(merge_settings_frame, style='Custom.TFrame')
        merge_mode_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(merge_mode_frame, text="合并方式：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(merge_mode_frame, text="合并到单个Sheet", variable=self.merge_mode, 
                       value="single", command=self.on_merge_mode_change, style='Custom.TRadiobutton').pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(merge_mode_frame, text="每个文件单独一个Sheet", variable=self.merge_mode, 
                       value="multiple", command=self.on_merge_mode_change, style='Custom.TRadiobutton').pack(side=tk.LEFT, padx=20)
        
        # Sheet名称设置（只在单sheet模式下显示）
        self.single_sheet_frame = ttk.Frame(merge_settings_frame, style='Custom.TFrame')
        ttk.Label(self.single_sheet_frame, text="Sheet名称：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(self.single_sheet_frame, textvariable=self.custom_sheet_name, width=30, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        
        # 多sheet模式下的设置
        self.multiple_sheet_frame = ttk.Frame(merge_settings_frame, style='Custom.TFrame')
        ttk.Label(self.multiple_sheet_frame, text="Sheet命名方式：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(self.multiple_sheet_frame, text="使用文件名", variable=self.sheet_name_mode, 
                       value="auto", command=self.on_sheet_name_mode_change, style='Custom.TRadiobutton').pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(self.multiple_sheet_frame, text="使用原Sheet名", variable=self.sheet_name_mode, 
                       value="original", command=self.on_sheet_name_mode_change, style='Custom.TRadiobutton').pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(self.multiple_sheet_frame, text="使用自定义名称", variable=self.sheet_name_mode, 
                       value="custom", command=self.on_sheet_name_mode_change, style='Custom.TRadiobutton').pack(side=tk.LEFT, padx=20)

        # 根据当前模式显示相应的frame
        self.on_merge_mode_change()

        # 数据区间选择
        range_frame = ttk.LabelFrame(merge_page, text="数据区间选择", padding="10", style='Custom.TLabelframe')
        range_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 行设置
        row_frame = ttk.Frame(range_frame, style='Custom.TFrame')
        row_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(row_frame, text="起始行：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(row_frame, textvariable=self.start_row, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        ttk.Label(row_frame, text="结束行：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(row_frame, textvariable=self.end_row, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        
        # 列设置
        col_frame = ttk.Frame(range_frame, style='Custom.TFrame')
        col_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(col_frame, text="起始列：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(col_frame, textvariable=self.start_col, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        ttk.Label(col_frame, text="结束列：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(col_frame, textvariable=self.end_col, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        
        ttk.Label(range_frame, text="注：列请使用Excel列标（如：A、B、C...）", style='Custom.TLabel').pack(padx=5, pady=5)
        
        # 表头设置
        header_frame = ttk.LabelFrame(merge_page, text="表头设置", padding="10", style='Custom.TLabelframe')
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        header_settings = ttk.Frame(header_frame, style='Custom.TFrame')
        header_settings.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(header_settings, text="表头行号：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(header_settings, textvariable=self.header_row, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        ttk.Label(header_settings, text="（第几行是表头，从1开始）", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(header_settings, text="保留表头", variable=self.keep_header, style='Custom.TCheckbutton').pack(side=tk.LEFT, padx=20)
        
        # ===== 样式设置页面 =====
        style_options_frame = ttk.LabelFrame(style_page, text="样式选项", padding="10", style='Custom.TLabelframe')
        style_options_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 添加主题选择
        theme_frame = ttk.Frame(style_options_frame, style='Custom.TFrame')
        theme_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(theme_frame, text="主题选择：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        # 获取所有可用主题
        available_themes = sorted(self.style.get_themes())
        self.theme_combobox = ttk.Combobox(theme_frame, values=available_themes, width=20, style='Custom.TCombobox')
        self.theme_combobox.set(self.style.theme_use())  # 设置当前主题
        self.theme_combobox.pack(side=tk.LEFT, padx=5)
        self.theme_combobox.bind('<<ComboboxSelected>>', self.on_theme_changed)
        
        # 样式选项
        style_frame = ttk.Frame(style_options_frame, style='Custom.TFrame')
        style_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Checkbutton(style_frame, text="保留样式", variable=self.keep_styles, style='Custom.TCheckbutton').pack(side=tk.LEFT, padx=20)
        ttk.Checkbutton(style_frame, text="保留列宽", variable=self.keep_column_width, style='Custom.TCheckbutton').pack(side=tk.LEFT, padx=20)
        ttk.Checkbutton(style_frame, text="保留单元格格式", variable=self.keep_cell_format, style='Custom.TCheckbutton').pack(side=tk.LEFT, padx=20)
        ttk.Checkbutton(style_frame, text="保留颜色", variable=self.keep_colors, style='Custom.TCheckbutton').pack(side=tk.LEFT, padx=20)
        
        # ===== 定时任务页面 =====
        schedule_settings_frame = ttk.LabelFrame(schedule_page, text="定时设置", padding="10", style='Custom.TLabelframe')
        schedule_settings_frame.pack(fill=tk.X, padx=10, pady=5)
        
        schedule_time_frame = ttk.Frame(schedule_settings_frame, style='Custom.TFrame')
        schedule_time_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(schedule_time_frame, text="设置定时执行时间（24小时制）：", style='Custom.TLabel').pack(side=tk.LEFT, padx=5)
        ttk.Entry(schedule_time_frame, textvariable=self.time_var, width=10, style='Custom.TEntry').pack(side=tk.LEFT, padx=5)
        self.schedule_button = ttk.Button(schedule_time_frame, text="启动定时任务", command=self.toggle_schedule, style='Custom.TButton')
        self.schedule_button.pack(side=tk.LEFT, padx=20)
        
        # ===== 底部按钮和状态栏 =====
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 预览按钮
        preview_frame = ttk.Frame(bottom_frame)
        preview_frame.pack(side=tk.LEFT)
        ttk.Button(preview_frame, text="预览选中文件", command=self.preview_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(preview_frame, text="预览合并结果", command=self.preview_merged_data).pack(side=tk.LEFT, padx=5)
        
        # 执行按钮
        ttk.Button(bottom_frame, text="立即执行合并", command=self.merge_files).pack(side=tk.RIGHT, padx=5)
        
        # 状态栏
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(status_frame, textvariable=self.status_var, style='Custom.TLabel').pack(fill=tk.X)

    def change_sheet(self):
        """修改选中文件的Sheet"""
        selection = self.file_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择要修改的文件！")
            return
            
        item = selection[0]
        file_path = self.get_file_path_from_item(item)
        if file_path not in self.file_sheets:
            return
            
        # 创建Sheet选择窗口
        sheet_window = tk.Toplevel(self.root)
        sheet_window.title("选择Sheet")
        sheet_window.geometry("300x200")
        sheet_window.transient(self.root)  # 设置为主窗口的子窗口
        sheet_window.grab_set()  # 模态窗口
        
        # 创建Sheet列表
        sheet_list = tk.Listbox(sheet_window, width=40, height=10)
        sheet_list.pack(pady=10)
        
        # 添加Sheet选项
        for sheet in self.file_sheets[file_path]:
            sheet_list.insert(tk.END, sheet)
            
        # 如果已经选择了Sheet，选中它
        if file_path in self.selected_sheets:
            try:
                index = self.file_sheets[file_path].index(self.selected_sheets[file_path])
                sheet_list.selection_set(index)
            except ValueError:
                pass
                
        def confirm_selection():
            selection = sheet_list.curselection()
            if selection:
                selected_sheet = sheet_list.get(selection[0])
                self.selected_sheets[file_path] = selected_sheet
                self.file_tree.set(item, "选择Sheet", selected_sheet)
                # 如果当前是使用原sheet名模式，更新sheet名称
                if self.sheet_name_mode.get() == "original":
                    self.file_tree.set(item, "自定义Sheet名", selected_sheet)
                sheet_window.destroy()
            else:
                messagebox.showwarning("警告", "请先选择一个Sheet！")
                
        def cancel_selection():
            sheet_window.destroy()
        
        # 创建按钮框架
        btn_frame = ttk.Frame(sheet_window)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 添加确定和取消按钮
        ttk.Button(btn_frame, text="确定", command=confirm_selection).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=cancel_selection).pack(side=tk.RIGHT, padx=5)
        
        # 双击选择功能
        def on_double_click(event):
            confirm_selection()
            
        sheet_list.bind('<Double-Button-1>', on_double_click)
        
        # 设置窗口焦点并等待
        sheet_window.focus_set()
        self.root.wait_window(sheet_window)

    def get_file_path_from_item(self, item):
        """从树形视图项获取文件路径"""
        file_name = self.file_tree.item(item)['values'][0]
        for file_path in self.input_files:
            if os.path.basename(file_path) == file_name:
                return file_path
        return None

    def add_files(self):
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        for file in files:
            if file not in self.input_files:
                try:
                    # 读取文件的sheet列表
                    xl = pd.ExcelFile(file)
                    sheets = xl.sheet_names
                    
                    # 检查是否为空文件
                    if not sheets:
                        raise ValueError("文件不包含任何Sheet")
                        
                    self.file_sheets[file] = sheets
                    self.selected_sheets[file] = sheets[0]  # 默认选择第一个sheet
                    self.input_files.append(file)
                    
                    # 添加到树形视图
                    file_name = os.path.basename(file)
                    sheet_name = os.path.splitext(file_name)[0] if self.sheet_name_mode.get() == "auto" else ""
                    self.file_tree.insert("", tk.END, values=(file_name, sheets[0], sheet_name))
                    
                except Exception as e:
                    messagebox.showerror("错误", f"读取文件 {os.path.basename(file)} 时出错：{str(e)}")
                    continue
                    
        # 更新状态
        self.status_var.set(f"已添加 {len(self.input_files)} 个文件")
        
        # 如果是多sheet模式，更新sheet名称
        if self.merge_mode.get() == "multiple":
            self.on_sheet_name_mode_change()

    def clear_files(self):
        self.input_files = []
        self.file_sheets = {}
        self.selected_sheets = {}
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.status_var.set("就绪")

    def preview_data(self):
        """预览选中的文件"""
        selection = self.file_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择要预览的文件！")
            return
            
        file_path = self.get_file_path_from_item(selection[0])
        if not file_path or file_path not in self.selected_sheets:
            messagebox.showerror("错误", "请先为文件选择Sheet！")
            return
            
        try:
            self.show_preview_window(self.read_excel_range(file_path), f"预览: {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("错误", f"预览数据时出错：{str(e)}")

    def preview_merged_data(self):
        """预览合并后的数据"""
        if not self.input_files:
            messagebox.showerror("错误", "请先选择要合并的Excel文件！")
            return
            
        try:
            # 读取所有Excel文件的指定范围
            all_data = []
            for file in self.input_files:
                if file in self.selected_sheets:
                    df = self.read_excel_range(file)
                    if not df.empty:
                        # 如果是分sheet模式，不需要数据来源列
                        if self.merge_mode.get() == "multiple" and '数据来源' in df.columns:
                            df = df.drop(columns=['数据来源'])
                        all_data.append((file, df))
                        
            if not all_data:
                raise ValueError("没有有效的数据可以合并！")
                
            if self.merge_mode.get() == "single":
                # 检查表头一致性
                headers_consistent, message = self.check_headers_consistency([df for _, df in all_data])
                if not headers_consistent:
                    messagebox.showwarning("警告", f"发现表头不一致：\n{message}")
                    
                # 合并数据
                merged_df = self.smart_merge([df for _, df in all_data])
                self.show_preview_window(merged_df, "预览: 合并结果")
            else:
                # 创建多sheet预览窗口
                preview_window = tk.Toplevel(self.root)
                preview_window.title("预览: 多Sheet结果")
                
                # 获取屏幕尺寸
                screen_width = preview_window.winfo_screenwidth()
                screen_height = preview_window.winfo_screenheight()
                window_width = int(screen_width * 0.8)
                window_height = int(screen_height * 0.8)
                x = (screen_width - window_width) // 2
                y = (screen_height - window_height) // 2
                preview_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
                
                # 创建Notebook用于显示多个sheet
                notebook = ttk.Notebook(preview_window)
                notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
                
                # 为每个文件创建一个sheet页
                for file_path, df in all_data:
                    # 获取sheet名称
                    file_name = os.path.basename(file_path)
                    sheet_name = None
                    for item in self.file_tree.get_children():
                        if self.file_tree.item(item)['values'][0] == file_name:
                            sheet_name = self.file_tree.item(item)['values'][2]
                            break
                    
                    if not sheet_name:
                        sheet_name = os.path.splitext(file_name)[0]
                        
                    # 创建sheet页
                    sheet_frame = ttk.Frame(notebook)
                    notebook.add(sheet_frame, text=sheet_name)
                    
                    # 创建表格
                    tree = ttk.Treeview(sheet_frame)
                    
                    # 创建垂直滚动条
                    vsb = ttk.Scrollbar(sheet_frame, orient="vertical", command=tree.yview)
                    vsb.pack(side=tk.RIGHT, fill=tk.Y)
                    
                    # 创建水平滚动条
                    hsb = ttk.Scrollbar(sheet_frame, orient="horizontal", command=tree.xview)
                    hsb.pack(side=tk.BOTTOM, fill=tk.X)
                    
                    # 配置树形视图的滚动
                    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    
                    # 设置列
                    columns = list(df.columns)
                    tree["columns"] = columns
                    tree["show"] = "headings"  # 不显示第一个空列
                    
                    # 设置每列的标题和宽度
                    for col in columns:
                        tree.heading(col, text=str(col))
                        # 计算列宽度（根据数据内容）
                        max_width = max(
                            len(str(col)),  # 标题长度
                            df[col].astype(str).str.len().max() if len(df) > 0 else 0  # 数据长度
                        )
                        tree.column(col, width=min(max_width * 10, 300))  # 限制最大宽度为300像素
                    
                    # 添加数据行
                    for idx, row in df.iterrows():
                        if idx < 1000:  # 限制显示前1000行
                            values = [str(row[col]) for col in columns]
                            tree.insert("", tk.END, values=values)
                        else:
                            break
                    
                    # 添加数据统计信息
                    info_frame = ttk.Frame(sheet_frame)
                    info_frame.pack(fill=tk.X, padx=10, pady=5)
                    
                    # 显示总行数和当前显示行数
                    total_rows = len(df)
                    shown_rows = min(total_rows, 1000)
                    ttk.Label(info_frame, text=f"总行数: {total_rows}    显示行数: {shown_rows}").pack(side=tk.LEFT)
                    
                    # 如果数据被截断，显示提示信息
                    if total_rows > 1000:
                        ttk.Label(info_frame, text="（仅显示前1000行）", foreground="red").pack(side=tk.LEFT, padx=5)
                
                # 添加关闭按钮
                ttk.Button(preview_window, text="关闭", command=preview_window.destroy).pack(pady=5)
                
        except Exception as e:
            messagebox.showerror("错误", f"预览数据时出错：{str(e)}")

    def read_excel_range(self, file_path):
        """读取指定范围的Excel数据"""
        if file_path not in self.selected_sheets:
            return pd.DataFrame()
            
        selected_sheet = self.selected_sheets[file_path]
        
        try:
            # 获取表头行号
            header_row = int(self.header_row.get()) - 1  # 转换为0-based索引
            
            # 处理行列范围
            start_row = int(self.start_row.get()) - 1 if self.start_row.get() else None
            end_row = int(self.end_row.get()) if self.end_row.get() else None
            start_col = self.col_to_num(self.start_col.get()) if self.start_col.get() else None
            end_col = self.col_to_num(self.end_col.get()) + 1 if self.end_col.get() else None
            
            # 读取Excel文件，指定表头行
            df = pd.read_excel(file_path, sheet_name=selected_sheet, header=header_row)
            
            # 截取指定范围
            df = df.iloc[start_row:end_row, start_col:end_col]
            
            # 添加数据来源列
            df['数据来源'] = os.path.basename(file_path)
            
            return df
            
        except Exception as e:
            raise Exception(f"读取文件 {os.path.basename(file_path)} 的 {selected_sheet} 时出错: {str(e)}")

    def check_headers_consistency(self, dataframes):
        """检查所有数据框的表头是否一致"""
        if not dataframes:
            return False, "没有数据可供检查"
            
        # 获取第一个数据框的列（不包括'数据来源'列）
        base_columns = set(col for col in dataframes[0].columns if col != '数据来源')
        
        # 检查其他数据框的列是否与第一个相同
        inconsistent_files = []
        for i, df in enumerate(dataframes[1:], 1):
            current_columns = set(col for col in df.columns if col != '数据来源')
            if current_columns != base_columns:
                file_name = df['数据来源'].iloc[0]
                diff_cols = base_columns.symmetric_difference(current_columns)
                inconsistent_files.append(f"文件 {file_name} 的列不一致，差异列：{', '.join(diff_cols)}")
                
        if inconsistent_files:
            return False, "\n".join(inconsistent_files)
        return True, "所有文件的表头一致"

    def merge_files(self):
        if not self.input_files:
            messagebox.showerror("错误", "请先选择要合并的Excel文件！")
            return
            
        if not self.output_path:
            messagebox.showerror("错误", "请选择输出路径！")
            return
            
        # 在合并之前检查sheet名称冲突
        if self.merge_mode.get() == "multiple":
            conflicts = self.check_sheet_name_conflicts()
            if conflicts:
                if not self.resolve_sheet_name_conflicts():
                    return  # 用户取消了冲突解决
            
        try:
            # 读取所有Excel文件的指定范围
            all_data = []
            first_file = True
            header_styles = None
            data_styles = None
            
            for file in self.input_files:
                if file in self.selected_sheets:
                    df = self.read_excel_range(file)
                    if not df.empty:
                        # 如果是分sheet模式，不需要数据来源列
                        if self.merge_mode.get() == "multiple" and '数据来源' in df.columns:
                            df = df.drop(columns=['数据来源'])
                        all_data.append((file, df))
                        
                        # 从第一个文件获取样式模板
                        if first_file:
                            try:
                                wb = load_workbook(file)
                                header_styles, data_styles = self.get_column_styles(
                                    wb, 
                                    self.selected_sheets[file]
                                )
                                first_file = False
                            except Exception as style_error:
                                print(f"获取样式时出错: {style_error}")
                                first_file = False
            
            if not all_data:
                raise ValueError("没有有效的数据可以合并！")
                
            # 生成输出文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.output_filename_var.get()}_{timestamp}.xlsx"
            output_file = os.path.join(self.output_path, filename)
            
            # 让用户确认或修改文件名
            output_file = filedialog.asksaveasfilename(
                initialdir=self.output_path,
                initialfile=filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="保存合并文件"
            )
            
            if not output_file:  # 用户取消了保存
                return
                
            # 更新输出路径
            self.output_path = os.path.dirname(output_file)
            self.output_path_var.set(self.output_path)
            
            # 添加到最近使用路径
            if self.output_path not in self.recent_paths:
                self.recent_paths.insert(0, self.output_path)
                if len(self.recent_paths) > 5:
                    self.recent_paths.pop()
                self.save_recent_paths()
                # 更新下拉菜单
                self.path_combobox['values'] = self.recent_paths + [p for p in self.default_paths if p not in self.recent_paths]
            
            # 根据合并方式处理数据
            if self.merge_mode.get() == "single":
                # 合并到单个sheet
                # 检查表头一致性
                headers_consistent, message = self.check_headers_consistency([df for _, df in all_data])
                if not headers_consistent:
                    if not messagebox.askyesno("警告", f"发现表头不一致：\n{message}\n是否继续合并？"):
                        return
                
                # 智能合并数据
                merged_df = self.smart_merge([df for _, df in all_data])
                
                # 确定sheet名称
                sheet_name = self.custom_sheet_name.get() if self.sheet_name_mode.get() == "custom" else "合并结果"
                
                # 保存合并后的文件
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 应用样式
                    if header_styles and data_styles and self.keep_styles.get():
                        wb = writer.book
                        self.apply_column_styles(wb, sheet_name, header_styles, data_styles)
                        
                        # 调整列宽
                        if self.keep_column_width.get():
                            ws = wb[sheet_name]
                            for column in ws.columns:
                                max_length = 0
                                try:
                                    column_letter = get_column_letter(column[0].column)
                                    for cell in column:
                                        try:
                                            if cell.value:
                                                max_length = max(max_length, len(str(cell.value)))
                                        except:
                                            continue
                                    adjusted_width = (max_length + 2)
                                    ws.column_dimensions[column_letter].width = adjusted_width
                                except:
                                    continue
            else:
                # 每个文件一个sheet
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    for file_path, df in all_data:
                        # 从文件树中获取对应的自定义sheet名称
                        file_name = os.path.basename(file_path)
                        sheet_name = None
                        for item in self.file_tree.get_children():
                            if self.file_tree.item(item)['values'][0] == file_name:
                                sheet_name = self.file_tree.item(item)['values'][2]
                                break
                        
                        if not sheet_name:
                            sheet_name = os.path.splitext(file_name)[0]
                        
                        # 确保sheet名称有效
                        sheet_name = self.sanitize_sheet_name(sheet_name)
                        
                        # 保存数据
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # 应用样式
                        if header_styles and data_styles and self.keep_styles.get():
                            wb = writer.book
                            self.apply_column_styles(wb, sheet_name, header_styles, data_styles)
                            
                            # 调整列宽
                            if self.keep_column_width.get():
                                ws = wb[sheet_name]
                                for column in ws.columns:
                                    max_length = 0
                                    try:
                                        column_letter = get_column_letter(column[0].column)
                                        for cell in column:
                                            try:
                                                if cell.value:
                                                    max_length = max(max_length, len(str(cell.value)))
                                            except:
                                                continue
                                        adjusted_width = (max_length + 2)
                                        ws.column_dimensions[column_letter].width = adjusted_width
                                    except:
                                        continue
            
            self.status_var.set(f"合并完成！输出文件：{output_file}")
            messagebox.showinfo("成功", f"文件合并完成！\n共合并了 {len(all_data)} 个文件的数据")
            
        except Exception as e:
            self.status_var.set(f"错误：{str(e)}")
            messagebox.showerror("错误", f"合并过程中出现错误：{str(e)}")

    def sanitize_sheet_name(self, sheet_name):
        """确保sheet名称有效"""
        # Excel的sheet名称限制：
        # 1. 长度不能超过31个字符
        # 2. 不能包含特殊字符: [ ] : * ? / \
        # 3. 不能为空
        
        # 移除非法字符
        invalid_chars = r'[]*?/\\'
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
            
        # 限制长度
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
            
        # 确保不为空
        if not sheet_name:
            sheet_name = "Sheet1"
            
        return sheet_name

    def smart_merge(self, dataframes):
        """智能合并数据框列表"""
        if not dataframes:
            return pd.DataFrame()
            
        # 获取所有列名（除了'数据来源'）
        all_columns = set()
        for df in dataframes:
            all_columns.update([col for col in df.columns if col != '数据来源'])
            
        # 确保所有数据框都有相同的列
        for df in dataframes:
            for col in all_columns:
                if col not in df.columns:
                    df[col] = None
                    
        # 根据是否保留表头选择合并方式
        if self.keep_header.get():
            # 保留表头的合并方式
            result_df = pd.concat(dataframes, ignore_index=True)
        else:
            # 不保留表头的合并方式（跳过第一个文件之后的表头行）
            result_df = dataframes[0].copy()  # 第一个文件完整保留
            header_row = int(self.header_row.get())  # 获取表头行号
            
            # 合并其他文件，跳过表头行
            for df in dataframes[1:]:
                # 跳过表头行
                data_without_header = df.iloc[header_row:]
                result_df = pd.concat([result_df, data_without_header], ignore_index=True)
        
        # 调整列顺序，确保'数据来源'列在最后
        if '数据来源' in result_df.columns:
            cols = [col for col in result_df.columns if col != '数据来源'] + ['数据来源']
            result_df = result_df[cols]
        
        return result_df

    def select_output_path(self):
        """选择输出路径"""
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.output_path = path
            self.output_path_var.set(path)
            # 添加到最近使用路径
            if path not in self.recent_paths:
                self.recent_paths.insert(0, path)
                if len(self.recent_paths) > 5:  # 保留最近5个路径
                    self.recent_paths.pop()
                self.save_recent_paths()
                # 更新下拉菜单
                self.path_combobox['values'] = self.recent_paths + [p for p in self.default_paths if p not in self.recent_paths]

    def on_path_selected(self, event):
        """当从下拉菜单选择路径时"""
        selected_path = self.path_combobox.get()
        self.output_path = selected_path
        self.output_path_var.set(selected_path)

    def load_recent_paths(self):
        """加载最近使用的路径"""
        try:
            if os.path.exists(self.recent_paths_file):
                with open(self.recent_paths_file, 'r', encoding='utf-8') as f:
                    paths = f.read().splitlines()
                return [p for p in paths if os.path.exists(p)]  # 只返回仍然存在的路径
        except Exception:
            pass
        return []

    def save_recent_paths(self):
        """保存最近使用的路径"""
        try:
            with open(self.recent_paths_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(self.recent_paths))
        except Exception:
            pass

    def toggle_schedule(self):
        if not self.is_scheduling:
            try:
                # 验证时间格式
                time_parts = self.time_var.get().split(":")
                if len(time_parts) != 2:
                    raise ValueError("时间格式错误")
                    
                hour = int(time_parts[0])
                minute = int(time_parts[1])
                
                if not (0 <= hour <= 23 and 0 <= minute <= 59):
                    raise ValueError("时间范围错误")
                    
                self.schedule_time = f"{hour:02d}:{minute:02d}"
                self.start_schedule()
                self.schedule_button.configure(text="停止定时任务")
                self.is_scheduling = True
                self.status_var.set(f"定时任务已启动，将在每天 {self.schedule_time} 执行")
                
            except ValueError as e:
                messagebox.showerror("错误", "请输入有效的时间格式（HH:MM）")
        else:
            self.stop_schedule()
            self.schedule_button.configure(text="启动定时任务")
            self.is_scheduling = False
            self.status_var.set("定时任务已停止")
            
    def start_schedule(self):
        schedule.clear()
        schedule.every().day.at(self.schedule_time).do(self.merge_files)
        self.schedule_thread = threading.Thread(target=self.run_schedule, daemon=True)
        self.schedule_thread.start()
        
    def stop_schedule(self):
        schedule.clear()
        
    def run_schedule(self):
        while self.is_scheduling:
            schedule.run_pending()
            time.sleep(30)

    def col_to_num(self, col_str):
        """将Excel列标转换为数字"""
        num = 0
        for c in col_str.upper():
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num - 1

    def show_preview_window(self, data, title):
        """显示预览窗口"""
        preview_window = tk.Toplevel(self.root)
        preview_window.title(title)
        preview_window.geometry("1000x600")
        
        # 创建主框架
        main_frame = ttk.Frame(preview_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建树形视图
        tree = ttk.Treeview(main_frame)
        
        # 创建垂直滚动条
        vsb = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 配置树形视图的滚动
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 设置列
        columns = list(data.columns)
        tree["columns"] = columns
        tree["show"] = "headings"  # 不显示第一个空列
        
        # 设置每列的标题和宽度
        for col in columns:
            tree.heading(col, text=str(col))
            # 计算列宽度（根据数据内容）
            max_width = max(
                len(str(col)),  # 标题长度
                data[col].astype(str).str.len().max() if len(data) > 0 else 0  # 数据长度
            )
            tree.column(col, width=min(max_width * 10, 300))  # 限制最大宽度为300像素
        
        # 添加数据行
        for idx, row in data.iterrows():
            if idx < 1000:  # 限制显示前1000行
                values = [str(row[col]) for col in columns]
                tree.insert("", tk.END, values=values)
            else:
                break
        
        # 添加数据统计信息
        info_frame = ttk.Frame(preview_window)
        info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 显示总行数和当前显示行数
        total_rows = len(data)
        shown_rows = min(total_rows, 1000)
        ttk.Label(info_frame, text=f"总行数: {total_rows}    显示行数: {shown_rows}").pack(side=tk.LEFT)
        
        # 如果数据被截断，显示提示信息
        if total_rows > 1000:
            ttk.Label(info_frame, text="（仅显示前1000行）", foreground="red").pack(side=tk.LEFT, padx=5)
            
        # 添加关闭按钮
        ttk.Button(preview_window, text="关闭", command=preview_window.destroy).pack(pady=5)

    def copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if not self.keep_styles.get():
            return

        if source_cell.has_style:
            # 复制字体样式
            if self.keep_styles.get():
                font_kwargs = {
                    'name': source_cell.font.name,
                    'size': source_cell.font.size,
                    'bold': source_cell.font.bold,
                    'italic': source_cell.font.italic,
                    'vertAlign': source_cell.font.vertAlign,
                    'underline': source_cell.font.underline,
                    'strike': source_cell.font.strike,
                }
                if self.keep_colors.get():
                    font_kwargs['color'] = source_cell.font.color
                target_cell.font = Font(**font_kwargs)
            
            # 复制填充样式（背景色）
            if self.keep_colors.get():
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # 复制边框样式
            if self.keep_styles.get():
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            
            # 复制对齐方式
            if self.keep_styles.get():
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    text_rotation=source_cell.alignment.text_rotation,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # 复制数字格式
            if self.keep_cell_format.get():
                target_cell.number_format = source_cell.number_format

    def get_column_styles(self, workbook, sheet_name):
        """获取列样式模板"""
        sheet = workbook[sheet_name]
        header_row = int(self.header_row.get())  # 使用用户指定的表头行
        
        # 获取表头行的样式
        header_styles = {}
        for cell in sheet[header_row]:
            if not isinstance(cell, type(None)):
                try:
                    col_letter = get_column_letter(cell.column)
                    if hasattr(cell, 'font'):
                        # 创建新的样式对象而不是直接引用
                        header_styles[col_letter] = {
                            'font': Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                vertAlign=cell.font.vertAlign,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color=cell.font.color
                            ),
                            'fill': PatternFill(
                                fill_type=cell.fill.fill_type,
                                start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color
                            ),
                            'border': Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            ),
                            'alignment': Alignment(
                                horizontal=cell.alignment.horizontal,
                                vertical=cell.alignment.vertical,
                                text_rotation=cell.alignment.text_rotation,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            ),
                            'number_format': cell.number_format
                        }
                except (AttributeError, TypeError):
                    continue
            
        # 获取数据行的样式（使用表头行下一行作为模板）
        data_styles = {}
        if sheet.max_row > header_row:
            for cell in sheet[header_row + 1]:
                if not isinstance(cell, type(None)):
                    try:
                        col_letter = get_column_letter(cell.column)
                        if hasattr(cell, 'font'):
                            # 创建新的样式对象而不是直接引用
                            data_styles[col_letter] = {
                                'font': Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                ),
                                'fill': PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color
                                ),
                                'border': Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                ),
                                'alignment': Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    text_rotation=cell.alignment.text_rotation,
                                    wrap_text=cell.alignment.wrap_text,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent
                                ),
                                'number_format': cell.number_format
                            }
                    except (AttributeError, TypeError):
                        continue
                
        return header_styles, data_styles

    def apply_column_styles(self, workbook, sheet_name, header_styles, data_styles):
        """应用列样式"""
        if not self.keep_styles.get():
            return
            
        sheet = workbook[sheet_name]
        
        try:
            # 应用表头样式
            for cell in sheet[1]:
                if not isinstance(cell, type(None)):
                    try:
                        col_letter = get_column_letter(cell.column)
                        if col_letter in header_styles:
                            style = header_styles[col_letter]
                            if self.keep_styles.get():
                                cell.font = style['font']
                                cell.border = style['border']
                                cell.alignment = style['alignment']
                            if self.keep_colors.get():
                                cell.fill = style['fill']
                            if self.keep_cell_format.get():
                                cell.number_format = style['number_format']
                    except (AttributeError, TypeError):
                        continue
                    
            # 应用数据行样式
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if not isinstance(cell, type(None)):
                        try:
                            col_letter = get_column_letter(cell.column)
                            if col_letter in data_styles:
                                style = data_styles[col_letter]
                                if self.keep_styles.get():
                                    cell.font = style['font']
                                    cell.border = style['border']
                                    cell.alignment = style['alignment']
                                if self.keep_colors.get():
                                    cell.fill = style['fill']
                                if self.keep_cell_format.get():
                                    cell.number_format = style['number_format']
                        except (AttributeError, TypeError):
                            continue
        except Exception as e:
            print(f"应用样式时出错: {str(e)}")
            # 继续执行，即使样式应用失败

    def get_custom_sheet_names(self, file_names):
        """获取用户自定义的sheet名称"""
        # 创建一个新窗口
        dialog = tk.Toplevel(self.root)
        dialog.title("设置Sheet名称")
        dialog.geometry("400x400")
        
        # 创建一个框架来容纳所有输入框
        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建一个Canvas和Scrollbar，用于处理多个文件的情况
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 存储所有Entry控件的字典
        entries = {}
        
        # 为每个文件创建一个输入框
        for i, file_name in enumerate(file_names):
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, padx=5, pady=2)
            
            ttk.Label(frame, text=f"文件: {file_name}").pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=20)
            entry.insert(0, os.path.splitext(file_name)[0])  # 默认使用文件名（不含扩展名）
            entry.pack(side=tk.RIGHT)
            entries[i] = entry
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 用于存储结果的变量
        result = {}
        
        def on_ok():
            result['names'] = [entry.get() for entry in entries.values()]
            dialog.destroy()
            
        def on_cancel():
            result['names'] = None
            dialog.destroy()
        
        # 添加确定和取消按钮
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT, padx=5)
        
        # 设置模态对话框
        dialog.transient(self.root)
        dialog.grab_set()
        self.root.wait_window(dialog)
        
        return result.get('names')

    def on_merge_mode_change(self):
        """当合并模式改变时的处理"""
        if self.merge_mode.get() == "single":
            self.multiple_sheet_frame.pack_forget()
            self.single_sheet_frame.pack(fill=tk.X, padx=5, pady=5)
            # 隐藏自定义Sheet名列
            self.file_tree.column("自定义Sheet名", width=0)
        else:
            self.single_sheet_frame.pack_forget()
            self.multiple_sheet_frame.pack(fill=tk.X, padx=5, pady=5)
            # 显示自定义Sheet名列
            self.file_tree.column("自定义Sheet名", width=200)
            self.on_sheet_name_mode_change()

    def on_sheet_name_mode_change(self):
        """当Sheet命名方式改变时的处理"""
        if self.merge_mode.get() == "multiple":
            for item in self.file_tree.get_children():
                file_name = self.file_tree.item(item)['values'][0]
                file_path = self.get_file_path_from_item(item)
                current_sheet = self.file_tree.item(item)['values'][1]  # 获取当前选中的sheet名
                
                if self.sheet_name_mode.get() == "auto":
                    # 使用文件名作为sheet名
                    sheet_name = os.path.splitext(file_name)[0]
                    self.file_tree.set(item, "自定义Sheet名", sheet_name)
                elif self.sheet_name_mode.get() == "original":
                    # 使用原sheet名
                    self.file_tree.set(item, "自定义Sheet名", current_sheet)
                else:
                    # 保持当前的自定义名称，如果没有则使用文件名
                    current_name = self.file_tree.item(item)['values'][2]
                    if not current_name:
                        sheet_name = os.path.splitext(file_name)[0]
                        self.file_tree.set(item, "自定义Sheet名", sheet_name)
            
            # 检查是否有名称冲突
            conflicts = self.check_sheet_name_conflicts()
            if conflicts:
                conflict_files = [f[1] for f in conflicts]
                messagebox.showwarning("警告", 
                    f"检测到以下文件的Sheet名称存在冲突：\n{', '.join(conflict_files)}\n"
                    "您可以：\n"
                    "1. 使用自定义名称模式手动修改sheet名称\n"
                    "2. 在保存时系统会提供冲突解决方案")

    def change_sheet_name(self):
        """修改选中文件的Sheet名称"""
        if self.merge_mode.get() != "multiple" or self.sheet_name_mode.get() != "custom":
            messagebox.showinfo("提示", "只有在分Sheet保存且使用自定义名称模式下才能修改Sheet名称")
            return
            
        selection = self.file_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择要修改的文件！")
            return
            
        item = selection[0]
        current_name = self.file_tree.item(item)['values'][2]
        
        # 创建输入对话框
        dialog = tk.Toplevel(self.root)
        dialog.title("修改Sheet名称")
        dialog.geometry("300x120")
        
        ttk.Label(dialog, text="请输入新的Sheet名称：").pack(padx=10, pady=5)
        entry = ttk.Entry(dialog, width=40)
        entry.insert(0, current_name)
        entry.pack(padx=10, pady=5)
        
        def on_ok():
            new_name = entry.get()
            if new_name:
                # 检查新名称是否会造成冲突
                existing_names = []
                for tree_item in self.file_tree.get_children():
                    if tree_item != item:  # 排除当前项
                        existing_names.append(self.file_tree.item(tree_item)['values'][2])
                
                if new_name in existing_names:
                    messagebox.showerror("错误", f"Sheet名称 '{new_name}' 已存在，请使用其他名称！")
                    return
                    
                self.file_tree.set(item, "自定义Sheet名", new_name)
                dialog.destroy()
            
        ttk.Button(dialog, text="确定", command=on_ok).pack(pady=10)
        
        # 设置模态对话框
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.focus_set()

    def check_sheet_name_conflicts(self):
        """检查sheet名称是否有冲突"""
        sheet_names = []
        conflicts = []
        
        for item in self.file_tree.get_children():
            sheet_name = self.file_tree.item(item)['values'][2]
            file_name = self.file_tree.item(item)['values'][0]
            if sheet_name in sheet_names:
                conflicts.append((sheet_name, file_name))
            else:
                sheet_names.append(sheet_name)
                
        return conflicts

    def suggest_sheet_name(self, base_name, existing_names):
        """为冲突的sheet名称生成建议名称"""
        counter = 1
        new_name = base_name
        while new_name in existing_names:
            new_name = f"{base_name}_{counter}"
            counter += 1
        return new_name

    def resolve_sheet_name_conflicts(self):
        """解决sheet名称冲突"""
        conflicts = self.check_sheet_name_conflicts()
        if not conflicts:
            return True
            
        # 创建冲突解决对话框
        dialog = tk.Toplevel(self.root)
        dialog.title("Sheet名称冲突")
        dialog.geometry("600x400")
        
        # 说明文本
        ttk.Label(dialog, text="检测到以下Sheet名称冲突，请选择处理方式：").pack(padx=10, pady=5)
        
        # 创建滚动框架
        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 创建滚动框架
        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 存储所有Entry控件
        entries = {}
        existing_names = [self.file_tree.item(item)['values'][2] for item in self.file_tree.get_children()]
        
        # 为每个冲突创建处理选项
        for sheet_name, file_name in conflicts:
            conflict_frame = ttk.LabelFrame(scrollable_frame, text=f"文件: {file_name}")
            conflict_frame.pack(fill=tk.X, padx=5, pady=5)
            
            ttk.Label(conflict_frame, text=f"当前Sheet名称: {sheet_name}").pack(padx=5, pady=2)
            
            # 建议的新名称
            suggested_name = self.suggest_sheet_name(sheet_name, existing_names)
            
            name_frame = ttk.Frame(conflict_frame)
            name_frame.pack(fill=tk.X, padx=5, pady=2)
            
            ttk.Label(name_frame, text="新名称：").pack(side=tk.LEFT)
            entry = ttk.Entry(name_frame, width=40)
            entry.insert(0, suggested_name)
            entry.pack(side=tk.LEFT, padx=5)
            
            # 使用建议按钮
            def use_suggestion(entry=entry, suggested=suggested_name):
                entry.delete(0, tk.END)
                entry.insert(0, suggested)
                
            ttk.Button(name_frame, text="使用建议名称", 
                      command=use_suggestion).pack(side=tk.LEFT, padx=5)
            
            entries[file_name] = entry
            existing_names.append(suggested_name)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 用于存储结果
        result = {"confirmed": False, "names": {}}
        
        def on_ok():
            # 检查新名称是否仍有冲突
            new_names = [entry.get() for entry in entries.values()]
            if len(new_names) != len(set(new_names)):
                messagebox.showerror("错误", "新的Sheet名称仍然存在冲突，请修改后重试！")
                return
                
            result["confirmed"] = True
            result["names"] = {file_name: entry.get() for file_name, entry in entries.items()}
            dialog.destroy()
            
        def on_cancel():
            result["confirmed"] = False
            dialog.destroy()
            
        # 按钮区域
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT, padx=5)
        
        # 设置模态对话框
        dialog.transient(self.root)
        dialog.grab_set()
        self.root.wait_window(dialog)
        
        # 如果用户确认了修改，更新文件树中的sheet名称
        if result["confirmed"]:
            for file_name, new_name in result["names"].items():
                for item in self.file_tree.get_children():
                    if self.file_tree.item(item)['values'][0] == file_name:
                        self.file_tree.set(item, "自定义Sheet名", new_name)
                        break
            return True
            
        return False

    def on_theme_changed(self, event):
        """当主题改变时的处理"""
        selected_theme = self.theme_combobox.get()
        self.style.set_theme(selected_theme)
        # 重新应用自定义样式
        self.configure_styles()
        # 强制更新底部按钮样式
        self.root.update_idletasks()
        # 更新状态栏提示
        self.status_var.set(f"已切换到主题：{selected_theme}")

def main():
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main() 