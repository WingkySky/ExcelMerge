"""
Excel合并核心逻辑模块
处理Excel文件的读取、合并等操作
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelMerger:
    def __init__(self, style_manager=None):
        """初始化Excel合并器"""
        self.style_manager = style_manager
        
    def merge_files(self, input_files, output_file, selected_sheets, file_sheets, merge_config):
        """
        执行Excel文件合并操作
        
        Args:
            input_files: 输入文件列表
            output_file: 输出文件路径
            selected_sheets: 选中的sheet信息 {文件路径: sheet名称}
            file_sheets: 文件的sheet信息 {文件路径: [sheet名称列表]}
            merge_config: 合并配置参数
            
        Returns:
            dict: 包含操作结果的字典
                - success: 是否成功
                - error: 错误信息（如果失败）
        """
        try:
            # 读取所有Excel文件的指定范围
            all_data = []
            first_file = True
            header_styles = None
            data_styles = None
            
            for file in input_files:
                if file in selected_sheets:
                    df = self.read_excel_range(
                        file,
                        selected_sheets[file],
                        merge_config['header_row'],
                        merge_config['start_row'],
                        merge_config['end_row'],
                        merge_config['start_col'],
                        merge_config['end_col'],
                        add_source=(merge_config['merge_mode'] == 'single')
                    )
                    
                    if not df.empty:
                        all_data.append((file, df))
                        
                        # 从第一个文件获取样式模板
                        if first_file and merge_config['keep_styles'] and self.style_manager:
                            try:
                                wb = load_workbook(file)
                                header_styles, data_styles, merged_cells = self.style_manager.get_column_styles(
                                    wb, 
                                    selected_sheets[file],
                                    merge_config['header_row']
                                )
                                first_file = False
                            except Exception as style_error:
                                print(f"获取样式时出错: {style_error}")
                                first_file = False
                                
            if not all_data:
                return {'success': False, 'error': "没有有效的数据可以合并！"}
                
            # 根据合并方式处理数据
            if merge_config['merge_mode'] == "single":
                # 检查表头一致性
                headers_consistent, message = self.check_headers_consistency([df for _, df in all_data])
                if not headers_consistent:
                    return {'success': False, 'error': f"表头不一致：\n{message}"}
                
                # 智能合并数据
                merged_df = self.smart_merge([df for _, df in all_data], merge_config['keep_header'])
                
                # 确定sheet名称
                sheet_name = merge_config['custom_sheet_name'] if merge_config['sheet_name_mode'] == "custom" else "合并结果"
                
                # 保存合并后的文件
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                    merged_df.to_excel(writer, sheet_name=sheet_name, index=False, float_format=None)
                    
                    # 应用样式
                    if merge_config['keep_styles'] and header_styles and data_styles and self.style_manager:
                        wb = writer.book
                        self.style_manager.apply_column_styles(
                            wb,
                            sheet_name,
                            header_styles,
                            data_styles,
                            merge_config,
                            merged_cells
                        )
                        
            else:
                # 每个文件一个sheet
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    for file_path, df in all_data:
                        # 获取sheet名称
                        file_name = os.path.basename(file_path)
                        if merge_config['sheet_name_mode'] == "auto":
                            sheet_name = os.path.splitext(file_name)[0]
                        elif merge_config['sheet_name_mode'] == "original":
                            sheet_name = selected_sheets[file_path]
                        else:  # custom
                            sheet_name = file_sheets[file_path].get('custom_name', os.path.splitext(file_name)[0])
                        
                        # 确保sheet名称有效
                        sheet_name = self.sanitize_sheet_name(sheet_name)
                        
                        # 保存数据
                        df.to_excel(writer, sheet_name=sheet_name, index=False, float_format=None)
                        
                        # 应用样式
                        if merge_config['keep_styles'] and header_styles and data_styles and self.style_manager:
                            wb = writer.book
                            self.style_manager.apply_column_styles(
                                wb,
                                sheet_name,
                                header_styles,
                                data_styles,
                                merge_config,
                                merged_cells
                            )
                            
            return {'success': True}
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
            
    def read_excel_range(self, file_path, sheet_name, header_row, start_row=None, end_row=None, 
                        start_col=None, end_col=None, add_source=True):
        """
        读取指定范围的Excel数据
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            header_row: 表头行号（1-based）
            start_row: 数据开始行号（1-based）
            end_row: 数据结束行号（1-based）
            start_col: 开始列（A, B, C...）
            end_col: 结束列（A, B, C...）
            add_source: 是否添加数据来源列
        """
        try:
            # 读取整个Excel文件，不指定表头
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # 处理列范围
            if start_col and str(start_col).strip():
                start_col_idx = self.col_to_num(start_col)
            else:
                start_col_idx = 0
                
            if end_col and str(end_col).strip():
                end_col_idx = self.col_to_num(end_col) + 1
            else:
                end_col_idx = len(df.columns)
            
            # 处理表头行
            header_row_idx = int(header_row) - 1 if header_row and str(header_row).strip() else 0
            
            # 获取表头数据
            header_df = df.iloc[header_row_idx:header_row_idx+1, start_col_idx:end_col_idx]
            header_values = header_df.iloc[0].values
            
            # 处理数据范围
            if not start_row or not str(start_row).strip():
                start_row_idx = header_row_idx + 1  # 默认从表头的下一行开始
            else:
                start_row_idx = int(start_row) - 1
            
            end_row_idx = int(end_row) if end_row and str(end_row).strip() else len(df)
            
            # 获取数据部分
            data_df = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx]
            
            # 设置列名
            columns = [str(val) if pd.notna(val) else f"Column_{i+1}" 
                      for i, val in enumerate(header_values)]
            data_df.columns = columns
            
            # 添加数据来源列
            if add_source:
                data_df['数据来源'] = os.path.basename(file_path)
            
            return data_df
            
        except Exception as e:
            raise Exception(f"读取文件 {os.path.basename(file_path)} 的 {sheet_name} 时出错: {str(e)}")
            
    def smart_merge(self, dataframes, keep_header=True):
        """智能合并数据框列表"""
        if not dataframes:
            return pd.DataFrame()
            
        # 获取所有列名（除了'数据来源'）
        all_columns = set()
        for df in dataframes:
            all_columns.update([col for col in df.columns if col != '数据来源'])
            
        # 确保所有数据框都有相同的列
        for df in dataframes:
            for col in all_columns:
                if col not in df.columns:
                    df[col] = None
                    
        # 根据是否保留表头选择合并方式
        if keep_header:
            # 保留表头的合并方式
            result_df = pd.concat(dataframes, ignore_index=True)
        else:
            # 不保留表头的合并方式（跳过第一个文件之后的表头行）
            result_df = dataframes[0].copy()  # 第一个文件完整保留
            
            # 合并其他文件，跳过表头行
            other_dfs = []
            for df in dataframes[1:]:
                # 确保列的顺序与第一个数据框一致
                df = df[result_df.columns]
                other_dfs.append(df)
            
            if other_dfs:
                result_df = pd.concat([result_df] + other_dfs, ignore_index=True)
        
        # 调整列顺序，确保'数据来源'列在最后
        if '数据来源' in result_df.columns:
            cols = [col for col in result_df.columns if col != '数据来源'] + ['数据来源']
            result_df = result_df[cols]
        
        return result_df
        
    def check_headers_consistency(self, dataframes):
        """检查所有数据框的表头是否一致"""
        if not dataframes:
            return False, "没有数据可供检查"
            
        # 获取第一个数据框的列（不包括'数据来源'列）
        base_columns = set(col for col in dataframes[0].columns if col != '数据来源')
        
        # 检查其他数据框的列是否与第一个相同
        inconsistent_files = []
        for i, df in enumerate(dataframes[1:], 1):
            current_columns = set(col for col in df.columns if col != '数据来源')
            if current_columns != base_columns:
                file_name = df['数据来源'].iloc[0]
                diff_cols = base_columns.symmetric_difference(current_columns)
                inconsistent_files.append(f"文件 {file_name} 的列不一致，差异列：{', '.join(diff_cols)}")
                
        if inconsistent_files:
            return False, "\n".join(inconsistent_files)
        return True, "所有文件的表头一致"
        
    @staticmethod
    def col_to_num(col_str):
        """将Excel列标转换为数字"""
        if not col_str:
            return None
        num = 0
        for c in col_str.upper():
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num - 1
        
    @staticmethod
    def sanitize_sheet_name(sheet_name):
        """确保sheet名称有效"""
        # Excel的sheet名称限制：
        # 1. 长度不能超过31个字符
        # 2. 不能包含特殊字符: [ ] : * ? / \
        # 3. 不能为空
        
        # 移除非法字符
        invalid_chars = r'[]*?/\\'
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
            
        # 限制长度
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
            
        # 确保不为空
        if not sheet_name:
            sheet_name = "Sheet1"
            
        return sheet_name 