"""
Excel样式管理模块
处理Excel文件样式的保存和应用
"""
from openpyxl.utils import get_column_letter
from copy import copy

class ExcelStyleManager:
    def __init__(self):
        """初始化样式管理器"""
        self.style_cache = {}  # 用于缓存样式
        
    def get_column_styles(self, workbook, sheet_name, header_row):
        """
        获取指定sheet的列样式
        
        Args:
            workbook: openpyxl的Workbook对象
            sheet_name: sheet名称
            header_row: 表头行号（1-based）
            
        Returns:
            tuple: (header_styles, data_styles, merged_cells)
        """
        try:
            sheet = workbook[sheet_name]
            header_styles = {}
            data_styles = {}
            
            # 获取合并单元格信息
            merged_cells = list(sheet.merged_cells.ranges)
            
            # 获取表头样式
            header_row_num = int(header_row)
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                cell = sheet[f"{col_letter}{header_row_num}"]
                header_styles[col] = {
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection),
                    'value': cell.value,  # 保存原始值，用于判断数据类型
                }
            
            # 获取数据区域样式（使用第一个数据行的样式作为模板）
            data_row_num = header_row_num + 1
            if data_row_num <= sheet.max_row:
                for col in range(1, sheet.max_column + 1):
                    col_letter = get_column_letter(col)
                    cell = sheet[f"{col_letter}{data_row_num}"]
                    data_styles[col] = {
                        'font': copy(cell.font),
                        'fill': copy(cell.fill),
                        'border': copy(cell.border),
                        'alignment': copy(cell.alignment),
                        'number_format': cell.number_format,
                        'protection': copy(cell.protection),
                        'value': cell.value,  # 保存原始值，用于判断数据类型
                    }
            
            return header_styles, data_styles, merged_cells
            
        except Exception as e:
            print(f"获取样式时出错: {str(e)}")
            return None, None, None
            
    def apply_column_styles(self, workbook, sheet_name, header_styles, data_styles, merge_config, merged_cells=None):
        """
        应用列样式到指定sheet
        
        Args:
            workbook: openpyxl的Workbook对象
            sheet_name: sheet名称
            header_styles: 表头样式字典
            data_styles: 数据样式字典
            merge_config: 合并配置
            merged_cells: 合并单元格信息
        """
        try:
            sheet = workbook[sheet_name]
            
            if merge_config['keep_styles']:
                # 应用合并单元格
                if merged_cells:
                    for merged_range in merged_cells:
                        # 调整合并范围以适应新的位置
                        sheet.merge_cells(str(merged_range))
                
                # 应用表头样式
                header_row = int(merge_config['header_row'])
                for col in range(1, sheet.max_column + 1):
                    if col in header_styles:
                        col_letter = get_column_letter(col)
                        cell = sheet[f"{col_letter}{header_row}"]
                        self._apply_cell_style(cell, header_styles[col])
                
                # 应用数据区域样式
                for row in range(header_row + 1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        if col in data_styles:
                            col_letter = get_column_letter(col)
                            cell = sheet[f"{col_letter}{row}"]
                            self._apply_cell_style(cell, data_styles[col])
                
                # 调整列宽
                self._adjust_column_width(sheet)
                    
        except Exception as e:
            print(f"应用样式时出错: {str(e)}")
            raise
            
    def _apply_cell_style(self, cell, style):
        """应用单元格样式"""
        try:
            # 根据原始值类型设置数字格式
            if isinstance(style.get('value'), (int, float)):
                cell.number_format = style['number_format']
            
            # 应用其他样式
            cell.font = copy(style['font'])
            cell.fill = copy(style['fill'])
            cell.border = copy(style['border'])
            cell.alignment = copy(style['alignment'])
            cell.protection = copy(style['protection'])
                    
        except Exception as e:
            print(f"应用单元格样式时出错: {str(e)}")
            
    def _adjust_column_width(self, sheet):
        """调整列宽"""
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for cell in sheet[col_letter]:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width 